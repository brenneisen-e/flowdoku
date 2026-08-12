# -*- coding: utf-8 -*-
"""
Erzeugt aus requirements_data.py beide Ergebnisse:

  docs/downloads/DEX-Requirements-kommentiert.xlsx   Formular für den Review
  docs/security-requirements.html                    verlinkte Matrix in der Doku

Aufruf aus dem Repository-Wurzelverzeichnis:
    python3 docs/tools/build_requirements.py

Die Vorlage („DEX - Requirements.xlsx“) bleibt unverändert: Spalten A bis E
werden übernommen wie sie sind, unsere Bewertung kommt in neue Spalten F bis H.
So bleibt erkennbar, was Vorgabe war und was wir geantwortet haben — und die
Spalte für den GTS Cyber Specialist bleibt leer, wie vorgesehen.
"""
import os
import sys
import html

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import requirements_data as R  # noqa: E402

import openpyxl  # noqa: E402
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side  # noqa: E402

ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
TEMPLATE = os.path.join(ROOT, 'docs', 'tools', 'requirements-template.xlsx')
XLSX_OUT = os.path.join(ROOT, 'docs', 'downloads', 'DEX-Requirements-kommentiert.xlsx')
HTML_OUT = os.path.join(ROOT, 'docs', 'security-requirements.html')

BY_ID = {r['id']: r for r in R.REQS}

CLS_FILL = {
    'dex': 'E6F2D5',
    'platform': 'E4EEF4',
    'na': 'EFEFEC',
    'open': 'FDF0DA',
    'flagged': 'FBE3E0',
}


def build_xlsx() -> None:
    wb = openpyxl.load_workbook(TEMPLATE)
    ws = wb['Sheet1']
    thin = Side(style='thin', color='C8CCC0')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = [
        (6, 'DEX — Einordnung'),
        (7, 'DEX — Wie die Anforderung erfüllt wird / warum sie nicht zutrifft'),
        (8, 'DEX — Offene Punkte, Abweichungen, Risiko'),
    ]
    for col, text in headers:
        c = ws.cell(row=5, column=col, value=text)
        c.font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor='4A5240')
        c.alignment = Alignment(wrap_text=True, vertical='top')
        c.border = border

    filled = 0
    for row in range(6, ws.max_row + 1):
        rid = ws.cell(row=row, column=1).value
        rid = str(rid).strip() if rid else ''
        if rid not in BY_ID:
            continue
        r = BY_ID[rid]
        label = R.CLS_LABEL[r['cls']][0]
        chapter = R.CHAPTERS.get(r['chapter'], '')

        ws.cell(row=row, column=6, value=label)
        ws.cell(row=row, column=7, value=r['met'])
        ws.cell(row=row, column=8, value=r.get('open', ''))

        # Spalte D (Evidence) ergaenzen statt ueberschreiben — der Hinweis der
        # Vorlage, dass der Assessor pruefen muss, bleibt stehen.
        prev = ws.cell(row=row, column=4).value
        prev = (str(prev).strip() + '\n\n') if prev else ''
        link = f'Systemarchitektur, Kapitel „{chapter}“ (docs/architektur.html#{r["chapter"]})'
        ws.cell(row=row, column=4, value=f'{prev}{r["evidence"]}\n{link}')

        for col in range(1, 9):
            c = ws.cell(row=row, column=col)
            c.alignment = Alignment(wrap_text=True, vertical='top')
            c.font = Font(name='Arial', size=10)
            c.border = border
        ws.cell(row=row, column=1).font = Font(name='Arial', size=10, bold=True)
        fill = PatternFill('solid', fgColor=CLS_FILL[r['cls']])
        for col in (1, 6):
            ws.cell(row=row, column=col).fill = fill
        ws.row_dimensions[row].height = None
        filled += 1

    widths = {'A': 13, 'B': 62, 'C': 40, 'D': 46, 'E': 20, 'F': 15, 'G': 62, 'H': 52}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = 'A6'

    # Deckblatt mit Legende und Zaehlung — was der Reviewer zuerst wissen will.
    if 'Legende' in wb.sheetnames:
        del wb['Legende']
    leg = wb.create_sheet('Legende', 0)
    leg['A1'] = 'DEX — System Security Requirements, kommentiert'
    leg['A1'].font = Font(name='Arial', size=14, bold=True)
    leg['A3'] = f'Stand der Anwendung: v{R.APP_VERSION}'
    leg['A4'] = f'Bewertet am: {R.ASSESSED_ON}'
    leg['A5'] = 'Ausführliche Begründung je Anforderung: docs/security-requirements.html'
    leg['A6'] = 'Architektur, auf die verwiesen wird: docs/architektur.html'
    for r_ in range(3, 7):
        leg.cell(row=r_, column=1).font = Font(name='Arial', size=10)

    leg['A8'] = 'Einordnung'
    leg['B8'] = 'Bedeutung'
    leg['C8'] = 'Anzahl'
    for col in ('A8', 'B8', 'C8'):
        leg[col].font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
        leg[col].fill = PatternFill('solid', fgColor='4A5240')

    counts = {}
    for r in R.REQS:
        counts[r['cls']] = counts.get(r['cls'], 0) + 1
    row_i = 9
    order = ['dex', 'platform', 'na', 'open', 'flagged']
    for key in order:
        short, long = R.CLS_LABEL[key]
        leg.cell(row=row_i, column=1, value=short).fill = PatternFill('solid', fgColor=CLS_FILL[key])
        leg.cell(row=row_i, column=2, value=long)
        leg.cell(row=row_i, column=3, value=counts.get(key, 0))
        for col in range(1, 4):
            leg.cell(row=row_i, column=col).font = Font(name='Arial', size=10)
        row_i += 1
    leg.cell(row=row_i, column=2, value='Summe').font = Font(name='Arial', size=10, bold=True)
    leg.cell(row=row_i, column=3, value=f'=SUM(C9:C{row_i - 1})').font = Font(name='Arial', size=10, bold=True)

    leg.cell(row=row_i + 2, column=1,
             value='Hinweis: Spalte E („Requirement Status“) bleibt bewusst leer — sie gehört dem '
                   'GTS Cyber Specialist. Die Spalten A bis E sind unverändert aus der Vorlage '
                   'übernommen; unsere Antworten stehen in F bis H, die Nachweis-Verweise sind in D '
                   'ergänzt.').font = Font(name='Arial', size=10, italic=True)
    leg.column_dimensions['A'].width = 18
    leg.column_dimensions['B'].width = 58
    leg.column_dimensions['C'].width = 10
    for r_ in range(1, row_i + 4):
        leg.cell(row=r_, column=1).alignment = Alignment(wrap_text=True, vertical='top')
        leg.cell(row=r_, column=2).alignment = Alignment(wrap_text=True, vertical='top')

    os.makedirs(os.path.dirname(XLSX_OUT), exist_ok=True)
    wb.save(XLSX_OUT)
    print(f'XLSX geschrieben: {XLSX_OUT} ({filled} Anforderungen befüllt)')


def esc(s: str) -> str:
    return html.escape(s or '').replace('\n', '<br>')


def build_html() -> None:
    counts = {}
    for r in R.REQS:
        counts[r['cls']] = counts.get(r['cls'], 0) + 1

    rows = []
    for r in R.REQS:
        short, long = R.CLS_LABEL[r['cls']]
        chapter = R.CHAPTERS.get(r['chapter'], '')
        openp = r.get('open', '')
        rows.append(f"""
      <article class="req" id="{r['id']}">
        <div class="req-head">
          <span class="req-id">{r['id']}</span>
          <span class="cls cls-{r['cls']}">{esc(short)}</span>
          <a class="req-chapter" href="architektur.html#{r['chapter']}">Kapitel &bdquo;{esc(chapter)}&ldquo;</a>
        </div>
        <p class="req-met">{esc(r['met'])}</p>
        <p class="req-ev"><span class="lbl">Nachweis</span>{esc(r['evidence'])}</p>
        {f'<p class="req-open"><span class="lbl">Offen</span>{esc(openp)}</p>' if openp else ''}
      </article>""")

    legend = ''.join(
        f'<div class="lg lg-{k}"><b>{counts.get(k, 0)}</b><span>{esc(R.CLS_LABEL[k][1])}</span></div>'
        for k in ['dex', 'platform', 'na', 'open', 'flagged'])

    doc = f"""<!DOCTYPE html>
<html lang="de">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<meta name="description" content="System Security Requirements für DEX — Bewertung je Anforderung mit Verweis in die Architekturdokumentation.">
<title>DEX Security Requirements</title>
<style>
  :root {{
    --bg:#f6f7f3; --surface:#fff; --surface-2:#edefe7; --ink:#191d16; --ink-2:#454e40;
    --muted:#6f7869; --line:#d8dccf; --accent:#5b8a12; --accent-ink:#3f6108;
    --font-text:"Segoe UI",system-ui,-apple-system,"Helvetica Neue",Arial,sans-serif;
    --font-mono:"Cascadia Mono",Consolas,"SF Mono",ui-monospace,Menlo,monospace;
    --c-dex:#5b8a12; --c-platform:#2c637f; --c-na:#6f7869; --c-open:#9c5a06; --c-flagged:#a3312a;
  }}
  @media (prefers-color-scheme: dark) {{
    :root:not([data-theme="light"]) {{
      --bg:#12140f; --surface:#1a1d16; --surface-2:#232820; --ink:#e9ece3; --ink-2:#bcc4b2;
      --muted:#8d9584; --line:#2f3529; --accent:#9fd23e; --accent-ink:#b6e05e;
      --c-dex:#9fd23e; --c-platform:#7fb8da; --c-na:#8d9584; --c-open:#e2a253; --c-flagged:#e8837a;
    }}
  }}
  :root[data-theme="dark"] {{
    --bg:#12140f; --surface:#1a1d16; --surface-2:#232820; --ink:#e9ece3; --ink-2:#bcc4b2;
    --muted:#8d9584; --line:#2f3529; --accent:#9fd23e; --accent-ink:#b6e05e;
    --c-dex:#9fd23e; --c-platform:#7fb8da; --c-na:#8d9584; --c-open:#e2a253; --c-flagged:#e8837a;
  }}
  *{{box-sizing:border-box}}
  body{{margin:0;background:var(--bg);color:var(--ink);font-family:var(--font-text);font-size:16px;line-height:1.6;-webkit-font-smoothing:antialiased}}
  .wrap{{max-width:1080px;margin:0 auto;padding:0 28px 96px}}
  header.mast{{border-bottom:1px solid var(--line);padding:56px 0 30px;margin-bottom:32px}}
  .eyebrow{{font-family:var(--font-mono);font-size:12px;letter-spacing:.14em;text-transform:uppercase;color:var(--muted);margin:0 0 14px}}
  h1{{font-size:clamp(1.9rem,4vw,2.8rem);line-height:1.07;letter-spacing:-.025em;margin:0 0 16px;text-wrap:balance}}
  .standfirst{{font-size:1.08rem;color:var(--ink-2);max-width:64ch;margin:0 0 24px}}
  .legend{{display:flex;flex-wrap:wrap;gap:0;border:1px solid var(--line);border-radius:4px;overflow:hidden;background:var(--surface)}}
  .lg{{flex:1 1 auto;min-width:150px;padding:12px 16px;border-right:1px solid var(--line);border-top:3px solid}}
  .lg:last-child{{border-right:0}}
  .lg b{{display:block;font-size:1.5rem;line-height:1.1;font-variant-numeric:tabular-nums}}
  .lg span{{font-size:.8rem;color:var(--ink-2)}}
  .lg-dex{{border-top-color:var(--c-dex)}} .lg-platform{{border-top-color:var(--c-platform)}}
  .lg-na{{border-top-color:var(--c-na)}} .lg-open{{border-top-color:var(--c-open)}}
  .lg-flagged{{border-top-color:var(--c-flagged)}}
  .intro{{max-width:74ch;margin:0 0 34px}}
  .intro p{{margin:0 0 12px}}
  .req{{border:1px solid var(--line);border-radius:4px;background:var(--surface);padding:16px 18px;margin-bottom:14px;scroll-margin-top:20px}}
  .req-head{{display:flex;flex-wrap:wrap;align-items:center;gap:12px;margin-bottom:10px}}
  .req-id{{font-family:var(--font-mono);font-weight:700;font-size:.95rem}}
  .cls{{font-family:var(--font-mono);font-size:10px;letter-spacing:.07em;text-transform:uppercase;padding:2px 7px;border-radius:3px;border:1px solid;white-space:nowrap}}
  .cls-dex{{color:var(--c-dex);border-color:var(--c-dex)}} .cls-platform{{color:var(--c-platform);border-color:var(--c-platform)}}
  .cls-na{{color:var(--c-na);border-color:var(--c-na)}} .cls-open{{color:var(--c-open);border-color:var(--c-open)}}
  .cls-flagged{{color:var(--c-flagged);border-color:var(--c-flagged)}}
  .req-chapter{{margin-left:auto;font-size:.82rem;color:var(--accent-ink);text-decoration:none;border-bottom:1px solid transparent}}
  .req-chapter:hover,.req-chapter:focus-visible{{border-bottom-color:currentColor}}
  .req-met{{margin:0 0 10px;max-width:80ch}}
  .req-ev,.req-open{{margin:0 0 6px;font-size:.88rem;color:var(--ink-2);max-width:80ch}}
  .req-open{{color:var(--c-open)}}
  .lbl{{display:inline-block;font-family:var(--font-mono);font-size:10px;letter-spacing:.08em;text-transform:uppercase;color:var(--muted);margin-right:8px}}
  .req-open .lbl{{color:var(--c-open)}}
  a{{color:var(--accent-ink)}}
  footer.end{{border-top:1px solid var(--line);margin-top:44px;padding-top:20px;font-size:.85rem;color:var(--muted);max-width:74ch}}
  @media (prefers-reduced-motion: reduce){{*{{animation:none!important;transition:none!important}}}}
</style>
</head>
<body>
<div class="wrap">
  <header class="mast">
    <p class="eyebrow">DEX Event Experience Platform · Stand v{R.APP_VERSION}</p>
    <h1>System Security Requirements — Antwort je Anforderung</h1>
    <p class="standfirst">
      49 Anforderungen, jede einzeln beantwortet und mit dem Kapitel der Architekturdokumentation
      verknüpft, das sie belegt. Was nicht zutrifft, ist begründet; was offen ist, steht als offen
      da — nicht als erfüllt.
    </p>
    <div class="legend">{legend}</div>
  </header>

  <div class="intro">
    <p>
      Der Zuschnitt dieser Anwendung bestimmt die meisten Antworten: DEX ist ein SharePoint-Webpart,
      das im Browser läuft. Es gibt keinen Server, keine Datenbank, kein eigenes Netz, keine eigene
      Anmeldung und keine eigene Sitzung. Anforderungen an Serverhärtung, Netztopologie oder
      Web-Application-Firewall gehen deshalb ins Leere; Anforderungen an Anmeldung, Verschlüsselung
      und Protokollierung erfüllt die Plattform, in der die Anwendung lebt.
    </p>
    <p>
      Bleibt das, was DEX selbst verantwortet: Rollen und Sichtbarkeiten, der Umgang mit
      personenbezogenen Daten, das Verhalten im Fehlerfall — und die Stellen, an denen wir bewusst
      abweichen. Die sind hier ausdrücklich benannt, mit Wirkung und Empfehlung, statt in einer
      Erfüllt-Meldung zu verschwinden.
    </p>
    <p>
      Dieselbe Bewertung liegt als Excel bei
      (<a href="downloads/DEX-Requirements-kommentiert.xlsx">DEX-Requirements-kommentiert.xlsx</a>);
      beide entstehen aus derselben Quelle und können nicht auseinanderlaufen.
    </p>
  </div>
{''.join(rows)}

  <footer class="end">
    Bewertet am {R.ASSESSED_ON} gegen den Stand v{R.APP_VERSION}. Aussagen zur Plattform beschreiben,
    was Microsoft 365 im Deloitte-Tenant leistet; ihren Nachweis führt der Plattformbetrieb, nicht
    dieses Projekt.
  </footer>
</div>
</body>
</html>
"""
    with open(HTML_OUT, 'w', encoding='utf-8') as f:
        f.write(doc)
    print(f'HTML geschrieben: {HTML_OUT} ({len(R.REQS)} Anforderungen)')


if __name__ == '__main__':
    build_xlsx()
    build_html()
